#!/usr/bin/env python3
"""
Random Forest predictor for log Kd (L/kg), seed=42, NO imputation.

This version uses the same engineered variables and exact feature order as the
manuscript-style RF model:
- has_aromatic_ring
- has_two_aromatics
- charge_state_anion
- charge_state_cation
- charge_state_neutral
- charge_state_zwitterionic
- surface area (m2/g)
- O wt%
- delta_PZCpH

Usage:
    python rf_predict_logKd_excel_inplace.py
    python rf_predict_logKd_excel_inplace.py --train
    python rf_predict_logKd_excel_inplace.py --training "cleaned_with_deltaPZCpH_no planar.xlsx" --input "Prediction_Input_Template.xlsx"

Artifacts created on training:
    - rf_model_seed42_no_impute.joblib
    - scaler_seed42_no_impute.joblib
    - feature_names_seed42_no_impute.joblib

Behavior:
    * Training:
      - loads the training Excel
      - engineers aromaticity and charge-state variables
      - drops rows with NaN/±inf ONLY (no filling)
      - fits StandardScaler on TRAIN DATA
      - trains RF (n_estimators=200, random_state=42)
      - saves model artifacts
    * Prediction:
      - reads the input Excel
      - engineers the same variables
      - checks required columns and numeric values
      - applies saved scaler/model
      - writes predictions to a new column
        'pred_log Kd (L/kg)' in the SAME file after making a .backup.xlsx copy
"""

import argparse
import shutil
import warnings
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler

# ===== Base paths =====
SCRIPT_DIR = Path(__file__).resolve().parent

# ===== Filenames =====
TRAINING_EXCEL = SCRIPT_DIR / "cleaned_with_deltaPZCpH_no planar.xlsx"
INPUT_EXCEL = SCRIPT_DIR / "Prediction_Input_Template.xlsx"
TARGET_COL = "log Kd (L/kg)"
PRED_COL = "pred_log Kd (L/kg)"

MODEL_PATH = SCRIPT_DIR / "rf_model_seed42_no_impute.joblib"
SCALER_PATH = SCRIPT_DIR / "scaler_seed42_no_impute.joblib"
FEATS_PATH = SCRIPT_DIR / "feature_names_seed42_no_impute.joblib"

SEED = 42

MOLECULAR_FEATURE_ORDER = [
    "log D (L/kg)",
    "molecular weight (g/mol)",
    "Log S (mol/L) at pH 7",
    "molecular volume  (cm3/mol/100)",
    "estimated log KOC  (L/kg)",
    "has_aromatic_ring",
    "has_two_aromatics",
    "charge_state_anion",
    "charge_state_cation",
    "charge_state_neutral",
    "charge_state_zwitterionic",
]
ADSORBENT_FEATURE_ORDER = [
    "surface area (m2/g)",
    "O wt%",
    "delta_PZCpH",
]
FEATURE_ORDER = MOLECULAR_FEATURE_ORDER + ADSORBENT_FEATURE_ORDER


# ===== Utilities =====
def _resolve_path(path_str: str) -> Path:
    path = Path(path_str)
    if not path.is_absolute():
        path = SCRIPT_DIR / path
    return path


def _read_excel_frame(excel_path: Path) -> pd.DataFrame:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path.resolve()}")
    return pd.read_excel(excel_path)


def _find_charge_columns(df: pd.DataFrame) -> tuple[str, str]:
    positive_candidates = [col for col in df.columns if "positive charge" in col.lower()]
    negative_candidates = [col for col in df.columns if "negative charge" in col.lower()]

    if not positive_candidates or not negative_candidates:
        raise ValueError(
            "Could not detect the positive and negative charge columns automatically. "
            "Please include columns such as 'positive charges' and 'negative charges'."
        )

    return positive_candidates[0], negative_candidates[0]


def _engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()

    required_raw = ["number of aromatic rings"]
    missing_raw = [col for col in required_raw if col not in work.columns]
    if missing_raw:
        raise ValueError(
            f"Missing required column(s) for engineered variables: {missing_raw}. "
            "The input file must contain these raw columns."
        )

    positive_col, negative_col = _find_charge_columns(work)

    work["number of aromatic rings"] = pd.to_numeric(work["number of aromatic rings"], errors="coerce")
    work[positive_col] = pd.to_numeric(work[positive_col], errors="coerce")
    work[negative_col] = pd.to_numeric(work[negative_col], errors="coerce")

    work["has_aromatic_ring"] = (work["number of aromatic rings"] >= 1).astype(float)
    work["has_two_aromatics"] = (work["number of aromatic rings"] >= 2).astype(float)

    work["charge_state_anion"] = ((work[positive_col] == 0) & (work[negative_col] == 1)).astype(float)
    work["charge_state_cation"] = ((work[positive_col] == 1) & (work[negative_col] == 0)).astype(float)
    work["charge_state_neutral"] = ((work[positive_col] == 0) & (work[negative_col] == 0)).astype(float)
    work["charge_state_zwitterionic"] = ((work[positive_col] == 1) & (work[negative_col] == 1)).astype(float)

    # Drop the raw columns that were converted into engineered variables
    work = work.drop(columns=[positive_col, negative_col, "number of aromatic rings"], errors="ignore")
    return work


def _make_training_Xy(df: pd.DataFrame):
    if TARGET_COL not in df.columns:
        raise ValueError(f"Target column '{TARGET_COL}' not found in training Excel.")

    work = df.replace([np.inf, -np.inf], np.nan).copy()
    work = _engineer_features(work)

    missing = [c for c in FEATURE_ORDER if c not in work.columns]
    if missing:
        raise ValueError(
            "Training Excel is missing required model features: "
            f"{missing}."
        )

    X = work[FEATURE_ORDER].copy()
    for c in FEATURE_ORDER:
        X[c] = pd.to_numeric(X[c], errors="coerce")
    y = pd.to_numeric(work[TARGET_COL], errors="coerce")

    data = pd.concat([X, y.rename(TARGET_COL)], axis=1).dropna(axis=0, how="any").copy()
    X = data.drop(columns=[TARGET_COL])
    y = data[TARGET_COL].astype(float)

    if X.empty:
        raise ValueError("No complete training rows remain after cleaning.")

    return X, y


def train_and_save(excel_path: Path):
    print("Training model from:", excel_path)
    df = _read_excel_frame(excel_path)
    X, y = _make_training_Xy(df)

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    rf = RandomForestRegressor(
        n_estimators=200,
        random_state=SEED,
        n_jobs=1,
    )
    rf.fit(X_scaled, y)

    joblib.dump(rf, MODEL_PATH)
    joblib.dump(scaler, SCALER_PATH)
    joblib.dump(list(X.columns), FEATS_PATH)
    print("Saved:", MODEL_PATH.name, SCALER_PATH.name, FEATS_PATH.name)
    return rf, scaler, list(X.columns)


def load_artifacts():
    rf = joblib.load(MODEL_PATH)
    scaler = joblib.load(SCALER_PATH)
    feats = joblib.load(FEATS_PATH)
    return rf, scaler, feats


def ensure_model(training_path: Path, force_retrain: bool = False):
    have_all = MODEL_PATH.exists() and SCALER_PATH.exists() and FEATS_PATH.exists()
    if force_retrain or not have_all:
        print("Training model...")
        return train_and_save(training_path)
    print("Loading existing model...")
    return load_artifacts()


# ===== Prediction helpers =====
def align_features(df_in: pd.DataFrame, required_cols):
    """
    Engineer variables, keep required_cols in order, coerce to numeric,
    and raise informative errors if values are missing/non-numeric.
    """
    work = df_in.replace([np.inf, -np.inf], np.nan).copy()
    work = _engineer_features(work)

    extra = [c for c in work.columns if c not in required_cols]
    if extra:
        warnings.warn(f"Ignoring extra columns not used by model: {extra}")

    missing = [c for c in required_cols if c not in work.columns]
    if missing:
        raise ValueError(
            f"Missing required columns after feature engineering: {missing}. "
            "Please check that the input file contains all needed raw descriptors."
        )

    X = work[required_cols].copy()
    for c in required_cols:
        X[c] = pd.to_numeric(X[c], errors="coerce")

    bad = X.isna()
    if bad.any().any():
        rows, cols = np.where(bad.values)
        details = []
        for r, cidx in zip(rows, cols):
            details.append(f"(row {r+2}, column '{required_cols[cidx]}')")
        raise ValueError(
            "Found non-numeric or missing values in the prediction input at: "
            + ", ".join(details)
            + ".\nNo imputation is performed—please correct the input cells."
        )

    return X


def predict(rf, scaler, X_df: pd.DataFrame):
    return rf.predict(scaler.transform(X_df))


def write_predictions_inplace(xlsx_path: Path, preds):
    wb = load_workbook(xlsx_path)
    ws = wb.worksheets[0]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if PRED_COL in headers:
        col_idx = headers.index(PRED_COL) + 1
    else:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=PRED_COL)

    for i, p in enumerate(preds, start=2):
        ws.cell(row=i, column=col_idx, value=float(p))

    wb.save(xlsx_path)


# ===== Main =====
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--train", action="store_true", help="Force retraining from the training Excel")
    parser.add_argument("--training", type=str, default=str(TRAINING_EXCEL), help="Path to training Excel")
    parser.add_argument("--input", type=str, default=str(INPUT_EXCEL), help="Path to prediction input Excel")
    args = parser.parse_args()

    training_path = _resolve_path(args.training)
    input_path = _resolve_path(args.input)

    rf, scaler, feature_names = ensure_model(training_path=training_path, force_retrain=args.train)

    if not input_path.exists():
        raise FileNotFoundError(f"Prediction input Excel not found: {input_path.resolve()}")

    df_in = _read_excel_frame(input_path)
    X = align_features(df_in, feature_names)
    preds = predict(rf, scaler, X)

    backup = input_path.with_suffix(".backup.xlsx")
    shutil.copy2(input_path, backup)
    print(f"Backup created: {backup.name}")

    write_predictions_inplace(input_path, preds)
    print(f"Predictions written into {input_path.name}, column '{PRED_COL}'")


if __name__ == "__main__":
    main()
